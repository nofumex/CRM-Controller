import csv
import glob
import json
import sys
import time
from collections import Counter
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import bot
from groq import Groq


NOT_REALIZED = "Не смог реализовать"
OUT_PATH = ROOT / "cold_leads_analysis.xlsx"
TRANSCRIPTS_PATH = ROOT / ".analysis_cache" / "groq_transcripts.json"
REASONED_ANALYSIS_PATH = ROOT / ".analysis_cache" / "groq_reasoned_analysis.json"


def load_json(path: Path) -> Dict[str, str]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, data: Dict[str, str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def latest_complete_csv() -> Path:
    candidates = sorted(glob.glob(str(ROOT / "reports" / "cold_leads_analysis_*.csv")))
    for item in reversed(candidates):
        with open(item, encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle, delimiter=";"))
        if len(rows) >= 180:
            return Path(item)
    raise RuntimeError("No complete cold lead CSV found in reports/")


def transcript_for_row(row: Dict[str, str], transcripts: Dict[str, str]) -> tuple[str, str, str]:
    note_ids = [item.strip() for item in (row.get("transcribed_call_note_ids") or "").split(",") if item.strip()]
    ok_ids: List[str] = []
    error_ids: List[str] = []
    parts: List[str] = []
    for note_id in note_ids:
        text = transcripts.get(note_id, "")
        if text and not text.startswith("[TRANSCRIPTION_ERROR]"):
            ok_ids.append(note_id)
            parts.append(f"Звонок {note_id}: {text}")
        elif text:
            error_ids.append(note_id)
    return ", ".join(ok_ids), ", ".join(error_ids), "\n\n".join(parts)


def analyze_with_reason(groq: Groq, row: Dict[str, str], transcript_text: str) -> str:
    reason = row.get("reason_not_realized") or "не указана"
    prompt = (
        "Ты анализируешь холодный звонок юридической компании по банкротству/долгам.\n"
        "Лид холодный: человек с просрочками по долгам, не оставлял явную заявку на разговор.\n"
        "Нужно объяснить, почему менеджер не смог реализовать холодный звонок.\n"
        "Обязательно учитывай ДВА источника: 1) CRM-причину, которую указал менеджер; 2) транскрипт звонка.\n"
        "Если CRM-причина противоречит транскрипту или выглядит неполной, прямо напиши это мягко.\n"
        "Пиши на русском, 3-5 предложений: контактность, открытие, выявление боли, работа с возражениями, следующий шаг.\n\n"
        f"ID сделки: {row.get('lead_id')}\n"
        f"Название: {row.get('lead_name')}\n"
        f"Текущий этап: {row.get('pipeline')} -> {row.get('status')}\n"
        f"CRM-причина НР: {reason}\n"
        f"Длительности звонков >15 сек: {row.get('call_durations_over_15s') or 'нет'}\n"
        f"Транскрипт:\n{transcript_text[:18000]}"
    )
    completion = groq.chat.completions.create(
        model="llama-3.1-8b-instant",
        messages=[
            {"role": "system", "content": "Ты строгий, практичный аналитик продаж. Отвечай конкретно, без воды."},
            {"role": "user", "content": prompt},
        ],
        temperature=0.2,
        max_tokens=900,
    )
    return completion.choices[0].message.content.strip()


def fallback_analysis(row: Dict[str, str]) -> str:
    reason = row.get("reason_not_realized") or ""
    calls_total = int(row.get("calls_total") or 0)
    long_calls = int(row.get("calls_over_15s") or 0)
    if long_calls == 0:
        base = "Нет звонка дольше 15 секунд, поэтому качество разговора по записи оценить нельзя."
    elif row.get("transcription_errors"):
        base = "Есть звонок дольше 15 секунд, но транскрипт не получен по технической причине."
    else:
        base = "Транскрипт отсутствует."
    if reason:
        return f"{base} CRM-причина менеджера: «{reason}». Вывод: сделка не была реализована по указанной причине, но без полноценного транскрипта нельзя проверить качество открытия, выявления боли и работы с возражениями."
    if calls_total == 0:
        return "В CRM нет звонков по контакту сделки. Нельзя оценить холодный звонок: коммуникация фактически не состоялась или не записана."
    return f"{base} CRM-причина НР не указана."


def main() -> None:
    source_csv = latest_complete_csv()
    with source_csv.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle, delimiter=";"))

    transcripts = load_json(TRANSCRIPTS_PATH)
    reasoned = load_json(REASONED_ANALYSIS_PATH)
    groq_key = bot.ENV.get("GROQ_API_KEY", "")
    groq = Groq(api_key=groq_key) if groq_key else None

    for index, row in enumerate(rows, start=1):
        ok_ids, error_ids, transcript_text = transcript_for_row(row, transcripts)
        row["transcribed_call_note_ids"] = ok_ids
        row["transcription_errors"] = error_ids
        row["transcript_excerpt"] = transcript_text[:5000] if transcript_text else row.get("transcript_excerpt", "")

        long_calls = int(row.get("calls_over_15s") or 0)
        cache_key = row["lead_id"]
        needs_groq = long_calls > 0 and transcript_text and groq is not None
        if needs_groq and cache_key not in reasoned:
            print(f"[{index}/{len(rows)}] reason-aware analysis {row['lead_id']}")
            reasoned[cache_key] = analyze_with_reason(groq, row, transcript_text)
            save_json(REASONED_ANALYSIS_PATH, reasoned)
            time.sleep(0.4)
        row["ai_analysis"] = reasoned.get(cache_key) if needs_groq else fallback_analysis(row)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    headers = [
        ("lead_id", "ID"),
        ("lead_name", "Название"),
        ("lead_url", "Ссылка"),
        ("pipeline", "Воронка"),
        ("status", "Этап"),
        ("reason_not_realized", "Причина НР из CRM"),
        ("ai_analysis", "AI-анализ с учетом причины и транскрипта"),
        ("calls_total", "Звонков всего"),
        ("calls_over_15s", "Звонков >15 сек"),
        ("call_durations_over_15s", "Длительности >15 сек"),
        ("transcribed_call_note_ids", "Расшифрованы note_id"),
        ("transcription_errors", "Ошибки транскрипта note_id"),
        ("responsible_user_id", "Ответственный ID"),
        ("transcript_excerpt", "Транскрипт / фрагмент"),
    ]
    ws.append([title for _, title in headers])
    for row in rows:
        ws.append([row.get(key, "") for key, _ in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for excel_row in ws.iter_rows(min_row=2):
        for cell in excel_row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    widths = {
        "A": 12,
        "B": 34,
        "C": 48,
        "D": 18,
        "E": 24,
        "F": 34,
        "G": 80,
        "H": 13,
        "I": 15,
        "J": 18,
        "K": 24,
        "L": 24,
        "M": 16,
        "N": 90,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    summary = wb.create_sheet("Summary")
    long_rows = [row for row in rows if int(row.get("calls_over_15s") or 0) > 0]
    success_transcripts = [row for row in rows if row.get("transcribed_call_note_ids")]
    not_realized = [row for row in rows if row.get("status") == NOT_REALIZED]
    summary_rows = [
        ("Метрика", "Значение"),
        ("Всего холодных сделок", len(rows)),
        ("Сделок со звонками", sum(int(row.get("calls_total") or 0) > 0 for row in rows)),
        ("Сделок со звонками >15 сек", len(long_rows)),
        ("Сделок с успешной транскрибацией", len(success_transcripts)),
        ("Сделок в Не смог реализовать", len(not_realized)),
        ("Исходный CSV", str(source_csv)),
    ]
    for item in summary_rows:
        summary.append(item)
    summary.append(())
    summary.append(("Этап", "Количество"))
    for key, count in Counter(f"{row.get('pipeline')} -> {row.get('status')}" for row in rows).most_common():
        summary.append((key, count))
    summary.append(())
    summary.append(("Причина НР из CRM", "Количество"))
    for key, count in Counter((row.get("reason_not_realized") or "(пусто)") for row in not_realized).most_common():
        summary.append((key, count))

    for row in summary.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    summary.column_dimensions["A"].width = 46
    summary.column_dimensions["B"].width = 90
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font

    wb.save(OUT_PATH)
    print(OUT_PATH)
    print(
        json.dumps(
            {
                "rows": len(rows),
                "with_calls_over_15s": len(long_rows),
                "with_successful_transcripts": len(success_transcripts),
                "not_realized": len(not_realized),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
